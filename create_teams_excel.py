#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
出場チーム管理用Excelファイル生成スクリプト
openpyxlライブラリを使用してExcelファイルを生成します
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxlライブラリがインストールされていません。")
    print("以下のコマンドでインストールしてください:")
    print("pip install openpyxl")
    exit(1)

# Excelブックを作成
wb = Workbook()
ws = wb.active
ws.title = "出場チーム"

# ヘッダー行を定義
headers = ["チーム名", "チーム代表者", "メンバー数", "参加予定", "連絡先", "メモ"]

# ヘッダーを追加
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    # ヘッダーのスタイル（背景色：青、文字色：白、太字）
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# サンプルデータを追加
sample_data = [
    ["サンプルチーム1", "田中太郎", 4, "〇", "090-1234-5678", ""],
    ["サンプルチーム2", "佐藤花子", 3, "〇", "090-9876-5432", ""],
]

for row_num, row_data in enumerate(sample_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="center")

# 列の幅を自動調整
column_widths = [20, 20, 12, 10, 18, 20]
for col_num, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + col_num)].width = width

# 枠線を追加
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=len(sample_data) + 1, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = thin_border

# ファイルを保存
output_path = "data/teams.xlsx"
wb.save(output_path)
print(f"✓ Excelファイルを作成しました: {output_path}")
